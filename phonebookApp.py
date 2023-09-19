import os
import csv
import openpyxl
import xlwt
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidgetItem, QHeaderView, QMenuBar, QMenu, QFileDialog, \
    QMessageBox, QInputDialog
import design
from PyQt5.QtGui import QTextCursor


# пишем класс, который наследуется от QMainWindow и от Ui_MainWindow(из нашего design.py)
class PhonebookApp(QMainWindow, design.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)  # для инициализации нашего дизайна
        self.tableWidget.setColumnWidth(0, 200)
        self.readFormat = ''
        self.pathToCsv = ''
        self.pathToExel = ''
        self.actionOpenCsv.triggered.connect(self.openCsv)  # обработчик нажатия на пункт меню "file - open csv"
        self.actionOpenExel.triggered.connect(self.openExel)  # обработчик нажатия на пункт меню "file - open exel"
        self.readCsvButton.clicked.connect(self.readCsv)  # обработчик нажатия на кнопку 'read csv file'
        self.readExelButton.clicked.connect(self.readExel)  # обработчик нажатия на кнопку 'read exel file'
        self.saveChangesButton.clicked.connect(self.saveChanges)  # обработчик нажатия на кнопку 'save changes'

        self.actionFindContact.triggered.connect(
            self.findContact)  # обработчик нажатия на пункт меню "contact - find contact"
        self.tableWidget.cellClicked.connect(self.cell_was_clicked)

    def setPathToCsv(self, pathToCsv):
        self.pathToCsv = pathToCsv

    def getPathToCsv(self):
        return self.pathToCsv

    def setPathToExel(self, pathToExel):
        self.pathToExel = pathToExel

    def getPathToExel(self):
        return self.pathToExel

    def setReadFormat(self, readFormat):
        self.readFormat = readFormat

    def getReadFormat(self):
        return self.readFormat

    def openCsv(self):
        print(f'был окрыт формат: {self.getReadFormat()}')
        pathToCsv = QFileDialog.getOpenFileName(self, "Open csv file", '*.csv')[0]
        self.setPathToCsv(pathToCsv)
        self.setReadFormat('csv')
        print(f'сейчас открыли формат: {self.getReadFormat()}')

    def readCsv(self):
        if self.getPathToCsv() != '':
            with open(self.getPathToCsv(), "r") as fileInput:
                s = 0  # индекс записываемой строки
                for row in csv.reader(fileInput):
                    self.tableWidget.setColumnCount(len(row))  # количество столбцов равно длине прочитанной строки
                    self.tableWidget.setRowCount(s + 1)  # количество отображаемых строк
                    item = QtWidgets.QTableWidgetItem()
                    self.tableWidget.setVerticalHeaderItem(s, item)
                    for i in range(0, len(row)):
                        self.tableWidget.setItem(s, i, QtWidgets.QTableWidgetItem(row[i]))
                    s += 1
        else:
            QMessageBox.warning(self, "Внимание", "сначала выберите csv файл для открытия", QMessageBox.Ok)

    def openExel(self):
        print(f'был окрыт формат: {self.getReadFormat()}')
        pathToExel = QFileDialog.getOpenFileName(self, "Open exel file", '*.xlsx')[0]
        self.setPathToExel(pathToExel)
        self.setReadFormat('exel')
        print(f'сейчас открыли формат: {self.getReadFormat()}')

    def readExel(self):
        if self.getPathToExel() != '':
            path = self.getPathToExel()
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active
            values = list(sheet.values)
            s = 0
            self.tableWidget.setRowCount(sheet.max_row)
            self.tableWidget.setColumnCount(sheet.max_column)
            # self.tableWidget.setHorizontalHeaderLabels(values[0])  # если есть шапка таблицы,  (не работает)
            for row in values:
                print(row)
                item = QtWidgets.QTableWidgetItem()
                self.tableWidget.setVerticalHeaderItem(s, item)
                for i in range(0, len(row)):
                    self.tableWidget.setItem(s, i, QtWidgets.QTableWidgetItem(str(row[i])))
                s += 1
        else:
            QMessageBox.warning(self, "Внимание", "сначала выберите xlsx файл для открытия", QMessageBox.Ok)

    def findContact(self):
        inputText, ok = QInputDialog.getText(self, 'find contact',
                                             'input name or phone number:')  # QInputDialog.getText() возвращает введенное значение и True/False
        if ok:
            if self.getPathToCsv() != '':
                with open(self.getPathToCsv(), "r") as fileInput:
                    s = 0  # индекс записываемой строки
                    for row in csv.reader(fileInput):
                        for el in row:
                            if inputText in el:
                                self.tableWidget.setRowCount(s + 1)  # количество отображаемых строк
                                item = QtWidgets.QTableWidgetItem()
                                self.tableWidget.setVerticalHeaderItem(s, item)
                                for i in range(0, len(row)):
                                    self.tableWidget.setItem(s, i, QtWidgets.QTableWidgetItem(row[i]))
                                s += 1
                                break
            else:
                QMessageBox.warning(self, "Внимание", "не выбран файл для открытия", QMessageBox.Ok)

    #функция считывает значение ячейки, на которую кликнули(пока не используем)
    def cell_was_clicked(self):
        rowNumber = self.tableWidget.currentRow()
        columnNumber = self.tableWidget.currentColumn()
        # print([rowNumber, columnNumber])
        # print(self.tableWidget.item(rowNumber, columnNumber).text())

    def saveChanges(self):
        if self.readFormat != '':
            rowsCount = self.tableWidget.rowCount()
            colsCount = self.tableWidget.columnCount()
            data = [] #список, куда будем помещать прочитанные из таблицы данные
            for row in range(rowsCount):
                tmp = [] #в этот список поместим данные из прочитанной строки таблицы
                for col in range(colsCount):
                    try:
                        tmp.append(self.tableWidget.item(row, col).text())
                    except:
                        tmp.append('')
                data.append(tmp)
            if self.readFormat == 'csv':
                if self.getPathToCsv() != '':
                    with open(self.getPathToCsv(), mode="w", newline='') as w_file:
                        writer = csv.writer(w_file)
                        for row in data:
                            writer.writerow(row)
                else:
                    QMessageBox.warning(self, "Внимание", "не выбран файл для открытия", QMessageBox.Ok)
            elif self.readFormat == 'exel':
                if self.getPathToExel() != '':
                    wb = openpyxl.Workbook()
                    sheet = wb.active
                    rowNumber = 1
                    for r in data:
                        for j in range(1, colsCount + 1):
                            val = sheet.cell(row=rowNumber, column=j)
                            val.value = r[j - 1]
                        rowNumber += 1
                    wb.save(self.getPathToExel())
                else:
                    QMessageBox.warning(self, "Внимание", "не выбран файл для открытия", QMessageBox.Ok)
